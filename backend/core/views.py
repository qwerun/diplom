from django.contrib.auth.models import User
from django.db.models import Count, Q, Sum
import mimetypes
from pathlib import Path

from django.http import FileResponse, Http404, HttpResponse
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import decorators, permissions, response, status, viewsets
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.views import APIView

from .integrations import MetricApiError, fetch_external_metrics
from .models import (
    Activity,
    ActivityMedia,
    ActivityResult,
    Campaign,
    Channel,
    MetricSource,
    MetricType,
    MetricValue,
    Report,
    Status,
    StatusTransition,
)
from .permissions import (
    CanEditActivityResult,
    CanEditMetrics,
    IsAdmin,
    IsAdminOrHead,
    IsAdminOrReadOnly,
    IsManagerOrReadOnly,
    IsManagerOrStatusOnly,
)
from .serializers import (
    ActivityResultSerializer,
    ActivityMediaSerializer,
    ActivitySerializer,
    CampaignSerializer,
    ChannelSerializer,
    MetricSourceSerializer,
    MetricTypeSerializer,
    MetricValueSerializer,
    ReportSerializer,
    StatusSerializer,
    StatusTransitionSerializer,
    UserSerializer,
)


class MeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        return response.Response(UserSerializer(request.user).data)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related("profile").all().order_by("username")
    serializer_class = UserSerializer
    permission_classes = [IsAdmin]
    search_fields = ["username", "first_name", "last_name", "email"]


class StatusViewSet(viewsets.ModelViewSet):
    queryset = Status.objects.all()
    serializer_class = StatusSerializer
    permission_classes = [IsAdminOrReadOnly]
    filterset_fields = ["entity_type"]


class StatusTransitionViewSet(viewsets.ModelViewSet):
    queryset = StatusTransition.objects.select_related("from_status", "to_status")
    serializer_class = StatusTransitionSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        entity_type = self.request.query_params.get("entity_type")
        if entity_type:
            queryset = queryset.filter(from_status__entity_type=entity_type)
        return queryset


class ChannelViewSet(viewsets.ModelViewSet):
    queryset = Channel.objects.all()
    serializer_class = ChannelSerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ["name"]


class MetricSourceViewSet(viewsets.ModelViewSet):
    queryset = MetricSource.objects.all()
    serializer_class = MetricSourceSerializer
    permission_classes = [IsAdminOrReadOnly]
    filterset_fields = ["type", "is_active"]


class CampaignViewSet(viewsets.ModelViewSet):
    queryset = Campaign.objects.select_related("responsible_user", "status")
    serializer_class = CampaignSerializer
    permission_classes = [IsManagerOrStatusOnly]
    filterset_fields = ["status", "responsible_user"]
    search_fields = ["name", "goal"]
    ordering_fields = ["start_date", "end_date", "budget"]


class ActivityViewSet(viewsets.ModelViewSet):
    queryset = Activity.objects.select_related(
        "campaign", "channel", "metric_source", "status"
    )
    serializer_class = ActivitySerializer
    permission_classes = [IsManagerOrStatusOnly]
    filterset_fields = ["campaign", "channel", "status", "metric_source"]
    search_fields = ["name", "description"]

    def get_permissions(self):
        if self.action == "collect_metrics":
            return [CanEditMetrics()]
        return super().get_permissions()

    @decorators.action(detail=True, methods=["post"], url_path="collect-metrics")
    def collect_metrics(self, request, pk=None):
        activity = self.get_object()
        # API-сбор разрешен только для источников типа API и только после того,
        # как исполнитель добавил ссылку на пост/результат активности.
        if activity.metric_source.type != MetricSource.TYPE_API:
            return response.Response(
                {"detail": "Автосбор доступен только для источников типа API."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        result = getattr(activity, "result", None)
        if not result or not result.result_url:
            return response.Response(
                {"detail": "Для API-сбора сначала добавьте результат: ссылку на пост."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        metric_values = list(activity.metric_values.select_related("metric_type"))
        if not metric_values:
            return response.Response(
                {"detail": "Сначала добавьте плановые метрики активности."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            api_metrics = fetch_external_metrics(activity.metric_source, result.result_url)
        except MetricApiError as error:
            return response.Response({"detail": str(error.detail[0])}, status=status.HTTP_400_BAD_REQUEST)

        updated_names = []
        skipped_names = []
        for metric in metric_values:
            metric_name = metric.metric_type.name.strip().lower()
            if metric_name not in api_metrics:
                skipped_names.append(metric.metric_type.name)
                continue
            metric.actual_value = Decimal(api_metrics[metric_name]).quantize(Decimal("0.01"))
            metric.save(update_fields=["actual_value", "collected_at"])
            updated_names.append(metric.metric_type.name)

        if not updated_names:
            return response.Response(
                {"detail": "API не вернул ни одну из метрик, добавленных к активности."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = MetricValueSerializer(metric_values, many=True)
        return response.Response(
            {
                "detail": f"Обновлено из {activity.metric_source.name}: {', '.join(updated_names)}.",
                "skipped": skipped_names,
                "metrics": serializer.data,
            }
        )


class ActivityResultViewSet(viewsets.ModelViewSet):
    queryset = ActivityResult.objects.select_related("activity")
    serializer_class = ActivityResultSerializer
    permission_classes = [CanEditActivityResult]
    filterset_fields = ["activity"]


class ActivityMediaViewSet(viewsets.ModelViewSet):
    queryset = ActivityMedia.objects.select_related("activity")
    serializer_class = ActivityMediaSerializer
    permission_classes = [IsManagerOrReadOnly]
    parser_classes = [MultiPartParser, FormParser]
    filterset_fields = ["activity"]

    def get_permissions(self):
        if self.action in {"preview", "download"}:
            return [permissions.AllowAny()]
        return super().get_permissions()

    def open_media_file(self):
        media = self.get_object()
        if not media.file:
            raise Http404("File is not attached.")
        try:
            return media, media.file.open("rb")
        except FileNotFoundError as error:
            raise Http404("File is not found on server storage.") from error

    @decorators.action(detail=True, methods=["get"])
    def preview(self, request, pk=None):
        media, file_handle = self.open_media_file()
        content_type = mimetypes.guess_type(media.file.name)[0] or "application/octet-stream"
        return FileResponse(file_handle, content_type=content_type)

    @decorators.action(detail=True, methods=["get"])
    def download(self, request, pk=None):
        media, file_handle = self.open_media_file()
        filename = media.title.strip() if media.title else Path(media.file.name).name
        if "." not in filename:
            original_extension = Path(media.file.name).suffix
            filename = f"{filename}{original_extension}"
        content_type = mimetypes.guess_type(media.file.name)[0] or "application/octet-stream"
        return FileResponse(file_handle, as_attachment=True, filename=filename, content_type=content_type)


class MetricTypeViewSet(viewsets.ModelViewSet):
    queryset = MetricType.objects.all()
    serializer_class = MetricTypeSerializer
    permission_classes = [IsAdminOrReadOnly]


class MetricValueViewSet(viewsets.ModelViewSet):
    queryset = MetricValue.objects.select_related("activity", "metric_type")
    serializer_class = MetricValueSerializer
    permission_classes = [CanEditMetrics]
    filterset_fields = ["activity", "metric_type"]


class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.select_related("campaign")
    serializer_class = ReportSerializer
    permission_classes = [IsAdminOrHead]
    filterset_fields = ["campaign"]

    @decorators.action(detail=False, methods=["post"])
    def generate(self, request):
        campaign_id = request.data.get("campaign")
        if not campaign_id:
            return response.Response(
                {"campaign": "Укажите кампанию."}, status=status.HTTP_400_BAD_REQUEST
            )
        report = Report.objects.create(campaign_id=campaign_id)
        report.file_path = f"/api/reports/{report.id}/xlsx/"
        report.save(update_fields=["file_path"])
        return response.Response(ReportSerializer(report).data, status=status.HTTP_201_CREATED)

    @decorators.action(detail=True, methods=["get"])
    def xlsx(self, request, pk=None):
        report = self.get_object()
        campaign = report.campaign
        activities = campaign.activities.select_related("channel", "status", "metric_source").prefetch_related(
            "metric_values__metric_type"
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Отчет"

        title_fill = PatternFill("solid", fgColor="F2A900")
        header_fill = PatternFill("solid", fgColor="263238")
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)

        sheet["A1"] = f"Отчет по кампании: {campaign.name}"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].fill = title_fill
        sheet.merge_cells("A1:H1")

        summary = [
            ("Цель", campaign.goal),
            ("Бюджет", float(campaign.budget)),
            ("Сроки", f"{campaign.start_date} - {campaign.end_date}"),
            ("Статус", campaign.status.name),
            ("Ответственный", campaign.responsible_user.get_full_name() or campaign.responsible_user.username),
        ]
        row = 3
        for label, value in summary:
            sheet.cell(row=row, column=1, value=label).font = bold_font
            sheet.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        sheet.cell(row=row, column=1, value="Активности").font = Font(size=13, bold=True)
        row += 1
        activity_headers = ["Название", "Канал", "Статус", "Источник", "Тип сбора", "Результат", "Комментарий"]
        for col, header in enumerate(activity_headers, start=1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1
        for activity in activities:
            result = getattr(activity, "result", None)
            values = [
                activity.name,
                activity.channel.name,
                activity.status.name,
                activity.metric_source.name,
                activity.metric_source.get_type_display(),
                result.result_url if result else "",
                result.comment if result else "",
            ]
            for col, value in enumerate(values, start=1):
                sheet.cell(row=row, column=col, value=value)
            row += 1

        row += 2
        sheet.cell(row=row, column=1, value="Метрики").font = Font(size=13, bold=True)
        row += 1
        metric_headers = ["Активность", "Метрика", "Ед. изм.", "План", "Факт", "Выполнение"]
        for col, header in enumerate(metric_headers, start=1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1
        for activity in activities:
            for metric in activity.metric_values.all():
                sheet.cell(row=row, column=1, value=activity.name)
                sheet.cell(row=row, column=2, value=metric.metric_type.name)
                sheet.cell(row=row, column=3, value=metric.metric_type.unit)
                sheet.cell(row=row, column=4, value=float(metric.planned_value))
                sheet.cell(row=row, column=5, value=float(metric.actual_value))
                sheet.cell(row=row, column=6, value=metric.completion_percent / 100)
                sheet.cell(row=row, column=6).number_format = "0.0%"
                row += 1

        for column in "ABCDEFGH":
            sheet.column_dimensions[column].width = 22
        for sheet_row in sheet.iter_rows():
            for cell in sheet_row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"campaign_report_{campaign.id}.xlsx"
        response_file = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response_file["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response_file


class AnalyticsViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        return response.Response(self.dashboard_data())

    @decorators.action(detail=False, methods=["get"])
    def dashboard(self, request):
        return response.Response(self.dashboard_data())

    @decorators.action(detail=False, methods=["get"])
    def channels(self, request):
        data = (
            Channel.objects.annotate(
                activities_count=Count("activities"),
                planned=Sum("activities__metric_values__planned_value"),
                actual=Sum("activities__metric_values__actual_value"),
            )
            .values("id", "name", "activities_count", "planned", "actual")
            .order_by("name")
        )
        return response.Response(list(data))

    @decorators.action(detail=False, methods=["get"], url_path="campaign/(?P<campaign_id>[^/.]+)")
    def campaign(self, request, campaign_id=None):
        values = (
            MetricValue.objects.filter(activity__campaign_id=campaign_id)
            .values("metric_type__name")
            .annotate(planned=Sum("planned_value"), actual=Sum("actual_value"))
            .order_by("metric_type__name")
        )
        return response.Response(list(values))

    def dashboard_data(self):
        campaign_statuses = (
            Status.objects.filter(entity_type=Status.ENTITY_CAMPAIGN)
            .annotate(count=Count("campaigns"))
            .values("name", "count")
        )
        activity_statuses = (
            Status.objects.filter(entity_type=Status.ENTITY_ACTIVITY)
            .annotate(count=Count("activities"))
            .values("name", "count")
        )
        return {
            "campaigns": Campaign.objects.count(),
            "active_campaigns": Campaign.objects.filter(status__name="Активна").count(),
            "finished_campaigns": Campaign.objects.filter(status__name="Завершена").count(),
            "activities": Activity.objects.count(),
            "budget_sum": Campaign.objects.aggregate(total=Sum("budget"))["total"] or 0,
            "campaign_statuses": list(campaign_statuses),
            "activity_statuses": list(activity_statuses),
            "plan_fact": list(
                MetricValue.objects.values("metric_type__name")
                .annotate(planned=Sum("planned_value"), actual=Sum("actual_value"))
                .filter(Q(planned__isnull=False) | Q(actual__isnull=False))
                .order_by("metric_type__name")
            ),
        }
